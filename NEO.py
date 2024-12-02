import tkinter as tk
from tkinter import messagebox
import pandas as pd
from datetime import datetime, timedelta
import sys
import os
from tkinter import ttk
from tkinterdnd2 import DND_FILES, TkinterDnD
import threading
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import ctypes

class ModernWin11GUI:
    def __init__(self, root):
        self.root = root
        self.root.title("NEO uNID Processor")
        
        try:

            self.root.tk.call('source', 'azure.tcl')
            self.root.tk.call('set_theme', 'dark')
            
            # Configure transparency
            self.root.attributes('-alpha', 0.97)  
            
            # Enable Windows 11 Mica effect
            DWMWA_USE_IMMERSIVE_DARK_MODE = 20
            DWMWA_MICA_EFFECT = 1029
            hwnd = ctypes.windll.user32.GetParent(self.root.winfo_id())
            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd, DWMWA_MICA_EFFECT,
                ctypes.byref(ctypes.c_int(1)), ctypes.sizeof(ctypes.c_int)
            )

            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd, DWMWA_USE_IMMERSIVE_DARK_MODE,
                ctypes.byref(ctypes.c_int(1)), ctypes.sizeof(ctypes.c_int)
            )
        except Exception:
            pass
        
        # Set window size and position
        window_width = 900  
        window_height = 650
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        center_x = int(screen_width/2 - window_width/2)
        center_y = int(screen_height/2 - window_height/2)
        self.root.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
        self.root.minsize(900, 650)
        
        # Configure modern style
        self.root.config(bg='#1c1c1c')  
        
        # Create main container with padding
        self.container = ttk.Frame(root)
        self.container.pack(fill=tk.BOTH, expand=True, padx=40, pady=30)
        
        header_frame = ttk.Frame(self.container, style='Glass.TFrame')
        header_frame.pack(fill=tk.X, pady=(0, 30))
        
        # App icon
        icon_label = ttk.Label(
            header_frame,
            text="📊",
            style="Glass.TLabel"
        )
        icon_label.pack(side=tk.LEFT, padx=(20, 15))
        
        # Title and subtitle
        title_frame = ttk.Frame(header_frame, style='Glass.TFrame')
        title_frame.pack(side=tk.LEFT, fill=tk.Y, padx=10)
        
        title_label = ttk.Label(
            title_frame,
            text="NEO uNID Processor",
            style="GlassTitle.TLabel"
        )
        title_label.pack(anchor=tk.W)
        
        subtitle_label = ttk.Label(
            title_frame,
            text="Drop your Excel file to process uNIDs",
            style="GlassSubtitle.TLabel"
        )
        subtitle_label.pack(anchor=tk.W)
        
        # Create modern drop zone frame with glass effect
        self.drop_frame = ttk.Frame(self.container, style="GlassCard.TFrame")
        self.drop_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Make the entire drop_frame a drop target
        self.drop_frame.drop_target_register(DND_FILES)
        self.drop_frame.dnd_bind('<<Drop>>', self.process_dropped_file)
        
        # Create inner content frame - centered in drop_frame
        self.content_frame = ttk.Frame(self.drop_frame, style="Glass.TFrame")
        self.content_frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
        
        # Add modern icon with glow effect
        icon_label = ttk.Label(
            self.content_frame,
            text="📄",
            style="GlassIcon.TLabel"
        )
        icon_label.pack(pady=(0, 25))
        
        self.drop_label = ttk.Label(
            self.content_frame,
            text="Drag Excel File Here",
            style="GlassDropZone.TLabel"
        )
        self.drop_label.pack()
        
        separator_label = ttk.Label(
            self.content_frame,
            text="or",
            style="GlassSeparator.TLabel"
        )
        separator_label.pack(pady=15)
        
        self.browse_button = ttk.Button(
            self.content_frame,
            text="Browse Files",
            style="Accent.TButton",
            command=self.process_clicked
        )
        self.browse_button.pack(pady=10)
        
        self.status_frame = ttk.Frame(self.container, style="GlassStatus.TFrame")
        self.status_frame.pack(fill=tk.X, pady=(20, 0))
        
        self.status_label = ttk.Label(
            self.status_frame,
            text="Ready to process files",
            style="GlassStatus.TLabel"
        )
        self.status_label.pack(side=tk.LEFT, padx=15, pady=8)
        
        self.configure_styles()
        self.bind_events()

    def configure_styles(self):
        style = ttk.Style()
        
        # Modern glass effect colors
        accent_color = '#60cdff'  
        glass_bg = '#2c2c2c'      
        glass_fg = '#ffffff'      
        glass_secondary = '#cccccc' 
        
        # Glass frame styles
        style.configure(
            "Glass.TFrame",
            background=glass_bg,
        )
        
        style.configure(
            "GlassCard.TFrame",
            background=glass_bg,
            borderwidth=1,
            relief="solid"
        )
        
        style.configure(
            "Glass.TLabel",
            font=('Segoe UI', 24),
            foreground=glass_fg,
            background=glass_bg
        )
        
        style.configure(
            "GlassTitle.TLabel",
            font=('Segoe UI', 28, 'bold'),
            foreground=glass_fg,
            background=glass_bg
        )
        
        style.configure(
            "GlassSubtitle.TLabel",
            font=('Segoe UI', 13),
            foreground=glass_secondary,
            background=glass_bg
        )
        
        style.configure(
            "GlassIcon.TLabel",
            font=('Segoe UI', 64),
            foreground=accent_color,
            background=glass_bg
        )
        
        style.configure(
            "GlassDropZone.TLabel",
            font=('Segoe UI', 18),
            foreground=glass_fg,
            background=glass_bg
        )
        
        style.configure(
            "GlassSeparator.TLabel",
            font=('Segoe UI', 13),
            foreground=glass_secondary,
            background=glass_bg
        )
        
        style.configure(
            "Accent.TButton",
            font=('Segoe UI', 12),
            padding=15
        )
        
        style.configure(
            "GlassStatus.TFrame",
            background=glass_bg,
            relief="solid",
            borderwidth=1
        )
        
        style.configure(
            "GlassStatus.TLabel",
            font=('Segoe UI', 11),
            foreground=glass_secondary,
            background=glass_bg
        )
        
        style.map("GlassCard.TFrame",
                 background=[('active', '#353535')],
                 relief=[('active', 'solid')],
                 borderwidth=[('active', 1)])
                 
        style.map("Accent.TButton",
                 background=[('active', accent_color)],
                 foreground=[('active', '#000000')])

    def bind_events(self):
        self.drop_frame.bind('<Enter>', self.on_hover_enter)
        self.drop_frame.bind('<Leave>', self.on_hover_leave)

    def on_hover_enter(self, event):
        """Handle mouse enter with glow effect"""
        self.drop_frame.configure(style="GlassCard.TFrame")
        self.drop_label.configure(foreground='#60cdff')  
        self.root.config(cursor="hand2")

    def on_hover_leave(self, event):
        """Handle mouse leave"""
        self.drop_frame.configure(style="GlassCard.TFrame")
        self.drop_label.configure(foreground='#ffffff')  
        self.root.config(cursor="")

    def get_previous_monday(self):
        """Get the date of the previous Monday in MMDDYY format"""
        today = datetime.now()
        days_since_monday = today.weekday()
        if days_since_monday == 0:
            previous_monday = today
        else:
            previous_monday = today - timedelta(days=days_since_monday)
        return previous_monday.strftime('%m%d%y')

    def validate_and_convert_unid(self, id_str):
        """Validate and convert an ID to uNID format"""
        id_str = str(id_str).strip()
        
        if not id_str:
            return ('', 'Empty ID')
        if not id_str.isdigit():
            return (id_str, 'Contains non-numeric characters')
        if len(id_str) != 8:
            return (id_str, f'Wrong length: {len(id_str)} digits (expected 8)')
        if not id_str.startswith('0'):
            return (id_str, 'Does not start with 0')
            
        return ('u' + id_str[1:], '')
    
    def format_excel(self, output_file, processed_ids):
        """Format the Excel file with instructions and styling"""
        # Create a new workbook using openpyxl
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active

        # Define styles
        center_aligned = Alignment(horizontal='center', vertical='center')
        red_font = Font(color='FF0000')
        red_bold_font = Font(color='FF0000', bold=True)
        bold_font = Font(bold=True)

        # Row 1: Instructions
        ws.merge_cells('A1:E1')
        ws['A1'] = "Instructions"
        ws['A1'].alignment = center_aligned
        ws['A1'].font = bold_font

        # Row 2: Email instructions
        ws.merge_cells('A2:E2')
        ws['A2'] = "Do you want to send emails? Answer '1' for Yes, and '0' for No in Cell F2"
        ws['A2'].alignment = center_aligned
        ws['A2'].font = red_font
        ws['F2'] = 0

        # Row 3: User ID instructions
        ws.merge_cells('A3:E3')
        ws['A3'] = "Enter user ID/Username/Email in the below column"
        ws['A3'].alignment = center_aligned
        ws['A3'].font = red_font

        # Row 4: Warning
        ws.merge_cells('A4:E4')
        ws['A4'] = "Do not remove instructions or change any headers"
        ws['A4'].alignment = center_aligned
        ws['A4'].font = red_font

        # Row 6: ID Header
        ws['A6'] = "ID"
        ws['A6'].font = red_bold_font

        # Add IDs starting from row 7
        for i, id_value in enumerate(processed_ids, start=7):
            ws[f'A{i}'] = id_value

        # Set all columns to width 8.43
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 9.14

        # Save the workbook
        wb.save(output_file)

    def process_file(self, file_path):
        """Process the Excel file"""
        try:
            # Update status
            self.status_label.config(text="Processing file...")
            self.root.update()
            
            # Generate output filenames
            date_str = self.get_previous_monday()
            output_dir = os.path.dirname(file_path)
            output_file = os.path.join(output_dir, f"NEO{date_str}.xlsx")
            error_file = os.path.join(output_dir, f"NEO{date_str}_errors.xlsx")
            
            # Read file
            df = pd.read_excel(file_path, dtype=str)
            column_data = df.iloc[:, 0]
            
            # Find section breaks
            section_breaks = column_data.str.contains('Rehire|Not Start', 
                                                    case=False, 
                                                    na=False)
            
            if any(section_breaks):
                first_break = section_breaks.idxmax()
                valid_data = column_data[:first_break]
            else:
                valid_data = column_data
                
            # Process data
            valid_data = valid_data[valid_data.str.lower() != 'unid']
            valid_data = valid_data.dropna()
            
            processed_ids = []
            error_records = []
            
            for idx, raw_id in enumerate(valid_data, 1):
                converted_id, error = self.validate_and_convert_unid(raw_id)
                if error:
                    error_records.append({
                        'Row': idx,
                        'Raw_ID': raw_id,
                        'Error': error
                    })
                else:
                    processed_ids.append(converted_id)
            
            # Sort the processed IDs
            processed_ids = sorted(list(set(processed_ids)))
            
            # Create and format the output file
            self.format_excel(output_file, processed_ids)
            
            # Save error report if needed
            if error_records:
                error_df = pd.DataFrame(error_records)
                error_df.to_excel(error_file, index=False)
                error_msg = (f"\n{len(error_records)} invalid IDs found!\n"
                           f"See error report: NEO{date_str}_errors.xlsx")
            else:
                error_msg = "\nNo errors found - all IDs were valid!"
            
            # Show success message
            messagebox.showinfo(
                "Success",
                f"Processing complete!\n"
                f"Found {len(processed_ids)} valid unique uNIDs."
                f"{error_msg}\n\n"
                f"Output saved as: NEO{date_str}.xlsx"
            )
            
            self.status_label.config(text="Ready to process file...")
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            self.status_label.config(text="Ready to process file...")


    def process_dropped_file(self, event):
        """Handle dropped file"""
        file_paths = self.root.tk.splitlist(event.data)
        if not file_paths:
            return
            
        file_path = file_paths[0]  # Take the first file if multiple are dropped
        file_path = file_path.replace('{', '').replace('}', '')  # Clean up the path
        
        if file_path.lower().endswith('.xlsx'):
            threading.Thread(target=self.process_file, args=(file_path,)).start()
        else:
            messagebox.showerror("Error", "Please drop an Excel (.xlsx) file")

    def process_clicked(self, event):
        """Handle click on drop zone"""
        from tkinter import filedialog
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx")]
        )
        if file_path:
            threading.Thread(target=self.process_file, args=(file_path,)).start()

def main():
    root = TkinterDnD.Tk()
    app = ModernWin11GUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()